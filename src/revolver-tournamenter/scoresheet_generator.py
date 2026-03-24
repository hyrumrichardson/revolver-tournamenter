import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# ------------------------------------------------------------
# CONFIG: Pairings (you can replace this with a file load)
# Each entry is a tuple: (Round, PlayerA, PlayerB)
# ------------------------------------------------------------
# pairings = [
#     ("Alice", "Bob"),
#     ("Charlie", "Diana"),
#     ("Alice", "Charlie"),
#     ("Bob", "Diana"),
#     ("Alice", "Diana"),
#     ("Bob", "Charlie")
# ]


# ------------------------------------------------------------
# HELPER: Extract unique players
# ------------------------------------------------------------
def get_players(pairings):
    players = set()
    for a, b in pairings:
        players.add(a)
        players.add(b)
    return sorted(list(players))

# ------------------------------------------------------------
# HELPER: Normalize formula for xlsx import
# ------------------------------------------------------------
def normalize_formula(formula: str) -> str:
    """
    Clean formulas so they import into Google Sheets correctly.
    Fixes parsing issues caused by openpyxl -> XLSX -> Google Sheets.
    """
    if formula is None:
        return ""
    
    # Remove leading/trailing whitespace
    formula = formula.strip()

    # Remove newlines / carriage returns
    formula = formula.replace("\n", "").replace("\r", "")

    # Replace unicode minus, smart quotes, nonbreaking spaces
    replacements = {
        "–": "-",   # en dash
        "—": "-",   # em dash
        "“": '"',
        "”": '"',
        "’": "'",
        "\u00A0": " ",  # non-breaking space
    }
    for bad, good in replacements.items():
        formula = formula.replace(bad, good)

    # Collapse multiple spaces
    while "  " in formula:
        formula = formula.replace("  ", " ")

    # Ensure starts with '='
    if not formula.startswith("="):
        formula = "=" + formula

    return formula

# ------------------------------------------------------------
# MAIN FUNCTION: Create tournament scoresheet
# ------------------------------------------------------------
def create_scoresheet(pairings, filename="tournament.xlsx"):
    wb = openpyxl.Workbook()
    
    # ------------------------------------------------------------
    # SHEET 1: MATCH ENTRY SHEET
    # ------------------------------------------------------------
    ws = wb.active
    ws.title = "Match Results"
    
    headers = [
        "Game No.", "Player A", "Player B",
        "A Score", "B Score",
    ]
    
    ws.append(headers)
    
    # Fill initial match list
    for index, pair in enumerate(pairings):
        ws.append([index+1, pair[0], pair[1], "", ""])
    
    # Auto-adjust column widths
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ------------------------------------------------------------
    # SHEET 1a: Pairings per player
    # ------------------------------------------------------------
    ps = wb.create_sheet("Pairings")

    players = get_players(pairings)

    pairing_headers = [
        "Player",
        "Opponent",
        "Game No.",
        "",
        "Player",
        "Opponent",
        "Game No.",
        "",
        "Player",
        "Opponent",
        "Game No.",
        "",
        "Player",
        "Opponent",
        "Game No.",
        ""
    ]

    ps.append(pairing_headers)

    plopppair = []
    for player in players:
        opponents = []

        for index, pairing in enumerate(pairings):
            if player == pairing[0]:
                opponents.append((index+1, pairing[1]))
            if player == pairing[1]:
                opponents.append((index+1, pairing[0]))
        
        plopppair.append([player, opponents])
    
    def pindex(i):
        if (i < len(plopppair)):
            return plopppair[i][0]
        else:
            return ''
    
    def oindex(i, j):
        if (i < len(plopppair)):
            if (j < len(plopppair[i][1])):
                return plopppair[i][0][j]
            else:
                return ''
        else:
            return ''
    
    entriesperrow = 4
    index = 0
    row = []
    while index<len(plopppair):
        for i in range(index, len(plopppair)):
            row.append(plopppair[i][0])
            row.append('')
            row.append('')
            row.append('')

            if i % entriesperrow == entriesperrow - 1:
                break

        ps.append(row)
        row = []
        jndex = 0
        jlen = len(plopppair[index][1])
        for j in range(jlen):
            for i in range(index, len(plopppair)):
                row.append('')
                row.append(plopppair[i][1][j][1])
                row.append(plopppair[i][1][j][0])
                row.append('')

                if i % entriesperrow == entriesperrow - 1:
                    break
            
            ps.append(row)
            row = []

        index += entriesperrow
        if (index >= len(plopppair)):
            break                     
        

    # ------------------------------------------------------------
    # SHEET 2: SCORESHEET
    # ------------------------------------------------------------
    ss = wb.create_sheet("Scores")

    players = get_players(pairings)

    score_headers = [
        "Player",
        "Match Wins",
        "Match Losses",
        "MW%",
        "OMW%",
        "Game Wins",
        "Game Losses",
        "GW%",
        "OGW%"
    ]
    
    ss.append(score_headers)

    # Insert players + formulas
    for i, player in enumerate(players, start=2):
        row = i

        # Wins
        wins_formula = f"""=COUNTIFS('Match Results'!$B$2:$B, A{i}, 'Match Results'!$D$2:$D,2) + COUNTIFS('Match Results'!$C$2:$C, A{i}, 'Match Results'!$E$2:$E,2) """

        # Losses
        losses_formula = f"""=COUNTIFS('Match Results'!$B$2:$B, A{i}, 'Match Results'!$D$2:$D,"<2") + COUNTIFS('Match Results'!$C$2:$C, A{i}, 'Match Results'!$E$2:$E,"<2") """

        # MW%
        mwp_formula = f"=IFERROR(ROUND(B{i}/(B{i}+C{i}),2),"")"

        # OMW%
        omwp_formula = f"""=LET(
            player, A{i},
            pairA, 'Match Results'!$B$2:$B, 
            pairB, 'Match Results'!$C$2:$C,

            opps, FLATTEN(IFERROR(FILTER(pairB, pairA = player),""), IFERROR(FILTER(pairA, pairB = player),"")),
            oppList, FILTER(opps, LEN(opps)),
            omwp, MAP(oppList,LAMBDA(opp,MAX(VLOOKUP(opp,$A$2:$D,4,FALSE),0.33))),
                
            ROUND(AVERAGE(omwp), 2)
            )"""

        # Game Wins
        gw_formula = f"=SUMIF('Match Results'!$B$2:$B, A{i},'Match Results'!$D$2:$D)+SUMIF('Match Results'!$C$2:$C,A{i},'Match Results'!$E$2:$E)"

        # Game Losses
        gl_formula = f"=SUMIF('Match Results'!$B$2:$B, A{i},'Match Results'!$E$2:$E)+SUMIF('Match Results'!$C$2:$C, A{i},'Match Results'!$D$2:$D)"

        # GW%
        gwp_formula = f"=IFERROR(ROUND(F{i}/(F{i}+G{i}),2),"")"
        
        # OGW%
        ogwp_formula = f"""=LET(
            player, A{i},
            pairA, 'Match Results'!$B$2:$B, 
            pairB, 'Match Results'!$C$2:$C,

            opps, FLATTEN(IFERROR(FILTER(pairB, pairA = player),""), IFERROR(FILTER(pairA, pairB = player),"")),
            oppList, FILTER(opps, LEN(opps)),
            omwp, MAP(oppList,LAMBDA(opp,MAX(VLOOKUP(opp,$A$2:$H,8,FALSE),0.33))),
                
            ROUND(AVERAGE(omwp), 2)
            )"""

        ss.append([
            player,
            normalize_formula(wins_formula),
            normalize_formula(losses_formula),
            normalize_formula(mwp_formula),
            normalize_formula(omwp_formula),
            normalize_formula(gw_formula),
            normalize_formula(gl_formula),
            normalize_formula(gwp_formula),
            normalize_formula(ogwp_formula)
        ])

    # Auto-adjust columns in Scores sheet
    for col in range(1, len(score_headers)+1):
        ss.column_dimensions[get_column_letter(col)].width = 16

    # ------------------------------------------------------------
    # SHEET 3: LEADERBOARD
    # ------------------------------------------------------------
    ls = wb.create_sheet("Leaderboard")

    players = get_players(pairings)

    lb_headers = [
        "Rank",
        "Player",
        "Match Wins",
        "OMW%",
        "GW%",
        "OGW%"
    ]
    
    ls.append(lb_headers)

    row = []
    for i,p in enumerate(players, start=1):
        row.append(i)
        if (i == 1):
            row.append("=SORT(Scores!A2:A,Scores!B2:B,FALSE,Scores!H2:H,FALSE,Scores!I2:I,FALSE)")
        else:
            row.append('')
        
        r = i+1
        row.append(f"""=IFERROR(VLOOKUP(B{r},Scores!$A$2:$I,2,FALSE),"")""")
        row.append(f"""=IFERROR(VLOOKUP(B{r},Scores!$A$2:$I,5,FALSE),"")""")
        row.append(f"""=IFERROR(VLOOKUP(B{r},Scores!$A$2:$I,8,FALSE),"")""")
        row.append(f"""=IFERROR(VLOOKUP(B{r},Scores!$A$2:$I,9,FALSE),"")""")

        ls.append(row)
        row = []

    # ------------------------------------------------------------
    # STYLING
    # ------------------------------------------------------------
    bold = Font(bold=True)
    for cell in ss[1]:
        cell.font = bold
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    for cell in ws[1]:
        cell.font = bold
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    for cell in ps[1]:
        cell.font = bold
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    
    for cell in ls[1]:
        cell.font = bold
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    # Save workbook
    wb.save(filename)
    print(f"Workbook successfully created: {filename}")

    return (wb, filename)


# ------------------------------------------------------------
# RUN
# ------------------------------------------------------------
# create_scoresheet(pairings)